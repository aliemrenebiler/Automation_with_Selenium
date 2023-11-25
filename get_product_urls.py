"""
This is an excel file generator for bathroom products.

- It gets the product codes from excel file,
- finds the product in multiple pages,
- and puts the product URLs to excel file.
"""

# pylint: disable=broad-except

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ----------------------------
# COFIGURATIONS
# ----------------------------
EXCEL_FILE_PATH = "/Users/aliemrenebiler/Desktop/vitra_urunler.xlsx"
SHEET_NAME = "Sheet1"
WEBSITES_ROW_NUMBER = 1
WEBSITES_START_COL_NUMBER = 3
PRODUCT_CODES_COL_NUMBER = 1
PRODUCTS_START_ROW_NUMBER = 2
WEBSITES = [
    {
        "url": "https://www.megabad.com",
        "query": lambda code: f"/search/?query={code}/#/q/{code}",
        "startup_close_element": (
            By.ID,
            "cmpbntyestxt",
        ),
        "product_element": (
            By.CLASS_NAME,
            "grid-item-name",
        ),
    },
    {
        "url": "https://victoriaplum.com",
        "query": lambda code: f"/search?query={code}",
        "product_element": (
            By.CLASS_NAME,
            "card__link",
        ),
    },
    {
        "url": "https://www.victorianplumbing.co.uk",
        "query": lambda code: f"/search?q={code}",
        "startup_close_element": (
            By.ID,
            "allbutton",
        ),
        "product_element": (
            By.CLASS_NAME,
            "ProductCardBottomstyles__ProductCardDescription-sc-1qihnsb-6",
        ),
    },
    {
        "url": "https://www.diy.com",
        "query": lambda code: f"/search?term={code}",
        "product_element": (
            By.CLASS_NAME,
            "_66091259",
        ),
    },
    {
        "url": "https://www.emero.de",
        "query": lambda code: f"/katalogsuche/?q={code}",
        "product_element": (
            By.CLASS_NAME,
            "c-product-tile__link",
        ),
    },
    {
        "url": "https://www.sonono.de",
        "query": lambda code: f"/katalogsuche/?q={code}",
        "product_element": (
            By.CLASS_NAME,
            "c-product-tile__link",
        ),
    },
    {
        "url": "https://www.castorama.fr",
        "query": lambda code: f"/search?term={code}",
        "startup_close_element": (
            By.ID,
            "truste-consent-button",
        ),
        "product_element": (
            By.XPATH,
            '//div[@data-test-id="product-panel"]/a',
        ),
    },
    {
        "url": "https://www.livea.fr",
        "query": lambda code: f"/recherche?search_query={code}",
        "startup_close_element": (
            By.NAME,
            "saveCookiesPlusPreferences",
        ),
        "product_element": (
            By.CLASS_NAME,
            "products__item",
        ),
    },
    {
        "url": "https://www.warmango.fr",
        "query": lambda code: f"/recherche?q={code}",
        "product_element": (
            By.ID,
            "product-link",
        ),
    },
    {
        "url": "https://www.bidet-shower.co.uk/",
        "query": lambda code: f"/search/?q={code}",
        "product_element": (
            By.XPATH,
            '//div[@class="product"]//a',
        ),
    },
    {
        "url": "https://www.ukbathrooms.com/",
        "query": lambda code: f"/shop/search.html?sq={code}",
        "product_element": (
            By.CLASS_NAME,
            "box--image__name",
        ),
    },
    {
        "url": "https://www.plumbnation.co.uk/",
        "query": lambda code: f"/search?q={code}",
        "product_element": (
            By.CLASS_NAME,
            "product-card-link",
        ),
    },
    {
        "url": "https://www.bathroomsuppliesonline.com/",
        "query": lambda code: f"/?search={code}",
        "product_element": (
            By.CLASS_NAME,
            "itemName",
        ),
    },
    {
        "url": "https://www.ergonomicdesigns.co.uk/",
        "query": lambda code: f"/search.html?qs={code}",
        "startup_close_element": (
            By.ID,
            "CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll",
        ),
        "product_element": (
            By.XPATH,
            '//div[@class="c-product"]//a',
        ),
    },
    {
        "url": "https://www.homesupply.co.uk/",
        "query": lambda code: f"/search.php?pg=1&stext={code}",
        "product_element": (
            By.XPATH,
            '//div[@class="product"]//a',
        ),
    },
    {
        "url": "https://www.amazon.it/",
        "query": lambda code: f"/s?k={code}",
        "product_element": (
            By.XPATH,
            '//div[@data-component-type="s-search-result"]//a',
        ),
    },
    {
        "url": "https://www.deghi.it/",
        "query": lambda code: f"/ricerca?term={code}&c_id=0",
        "startup_close_element": (
            By.ID,
            "CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll",
        ),
        "product_element": (
            By.XPATH,
            '//div[@class="product-wrap product-category cola-sm-10 cola-md-5 cola-xl-2"]/a',
        ),
    },
    # -----------------------------------------------
    # This website finds other matches too
    # {
    #     "url": "https://www.otto.de/",
    #     "query": lambda code: f"/suche/{code}",
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "find_tile__productLink",
    #     ),
    # },
    # -----------------------------------------------
    # This website finds other matches too
    # {
    #     "url": "https://www.ebay.de",
    #     "query": lambda code: f"/sch/i.html?&_nkw={code}",
    #     "product_element": (
    #         By.XPATH,
    #         "//li/div/div/a[@class='s-item__link']",
    #     ),
    # },
    # -----------------------------------------------
    # This website finds other matches too
    # {
    #     "url": "https://www.wayfair.de",
    #     "query": lambda code: f"/keyword.php?keyword={code}",
    # },
    # -----------------------------------------------
    # This website finds other matches too
    # {
    #     "url": "https://www.tradingdepot.co.uk/",
    #     "query": lambda code: f"/search/?q={code}",
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "",
    #     ),
    # },
    # -----------------------------------------------
    # It works, but if there is one product, it directs to the product page and cannot find it
    # {
    #     "url": "https://www.obadis.com",
    #     "query": lambda code: f"/de/catalogsearch/result/?q={code}",
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "product-item-link",
    #     ),
    # },
    # -----------------------------------------------
    # It works, but if there is one product, it directs to the product page and cannot find it
    # {
    #     "url": "https://www.skybad.de",
    #     "query": lambda code: f"/catalogsearch/result/?q={code}",
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "product-item-link",
    #     ),
    # },
    # -----------------------------------------------
    # It works, but if there is one product, it directs to the product page and cannot find it
    # {
    #     "url": "https://www.qssupplies.co.uk/",
    #     "query": lambda code: f"/searchresult.aspx?searchkey={code}",
    #     "product_element": (
    #         By.XPATH,
    #         '//div[@class="newgrid-prod-details"]//a',
    #     ),
    # },
    # -----------------------------------------------
    # Has a really hard captcha
    # {
    #     "url": "https://www.reuter.com/fr-fr",
    #     "query": lambda code: f"/catalogsearch/?q={code}",
    #     "startup_close_element": (
    #         By.ID,
    #         "econda-pp2-banner-accept-all-channels-button",
    #     ),
    #     "product_element": (
    #         By.XPATH,
    #         '//div[@id="js-product-listing"]/div/ul/li/a',
    #     ),
    # },
    # -----------------------------------------------
    # Somewhat it does not work
    # {
    #     "url": "https://www.batinea.com",
    #     "query": lambda code: f"/search/{code}",
    #     "product_element": (
    #         By.XPATH,
    #         '//div[@class="products"]/ol/li/div/div/a',
    #     ),
    # },
    # -----------------------------------------------
    # Somewhat it does not work, also slow
    # {
    #     "url": "https://www.rubberduckbathrooms.co.uk/",
    #     "query": lambda code: f"/products_search.php?search_string={code}",
    #     "product_element": (
    #         By.XPATH,
    #         '//h2[@class="product-item-title "]/a',
    #     ),
    # },
    # -----------------------------------------------
    # Too slow, need to log in
    # {
    #     "url": "https://www.showroomprive.com/",
    #     "query": lambda code: f"",
    #     "product_element": (
    #         By.ID,
    #         "agree_button",
    #     ),
    # },
    # -----------------------------------------------
    # Too slow
    # {
    #     "url": "https://www.manomano.co.uk/",
    #     "query": lambda code: f"/search/{code}",
    #     "startup_close_element": (
    #         By.ID,
    #         "didomi-notice-agree-button",
    #     ),
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "oFBZUY",
    #     ),
    # },
    # -----------------------------------------------
    # Too slow
    # {
    #     "url": "https://www.manomano.de/",
    #     "query": lambda code: f"/search/{code}",
    #     "startup_close_element": (
    #         By.ID,
    #         "didomi-notice-agree-button",
    #     ),
    #     "product_element": (
    #         By.CLASS_NAME,
    #         "oFBZUY",
    #     ),
    # },
    # -----------------------------------------------
]
TIMEOUT = 3
# ----------------------------


def close_popup(browser: webdriver, close_element):
    "Handles the cookie pop up for specified page."

    WebDriverWait(browser, TIMEOUT).until(
        EC.presence_of_element_located(close_element)
    ).click()


def create_query_url(url, query, code):
    "Creates the search query for the website."

    cleared_url = url if url[-1] != "/" else url[:-1]
    cleared_query = query(code) if query(code)[0] != "/" else query(code)[1:]

    return f"{cleared_url}/{cleared_query}"


def find_href_in_element(browser: webdriver, product_element):
    "Finds the specified html item with class name."

    return (
        WebDriverWait(browser, TIMEOUT)
        .until(EC.presence_of_element_located(product_element))
        .get_attribute("href")
    )


def main():
    "Main code"

    # Open excel file
    workbook = load_workbook(EXCEL_FILE_PATH)
    try:
        sheet = workbook[SHEET_NAME]
    except Exception:
        sheet = workbook.active

    # Get product codes
    product_codes = [
        x.value
        for x in sheet[get_column_letter(PRODUCT_CODES_COL_NUMBER)][
            PRODUCTS_START_ROW_NUMBER - 1 :
        ]
    ]

    # Create browser
    active_browser = webdriver.Chrome()

    for website_index, website in enumerate(WEBSITES):
        try:
            # Get main page
            active_browser.get(website["url"])

            # Close pop-up at start for the website
            if "startup_close_element" in website.keys():
                close_popup(active_browser, website["startup_close_element"])

            # Add website to excel file
            sheet[
                get_column_letter(website_index + WEBSITES_START_COL_NUMBER)
                + str(WEBSITES_ROW_NUMBER)
            ] = website["url"]

            # Save excel file
            workbook.save(EXCEL_FILE_PATH)

            for product_index, product_code in enumerate(product_codes):
                try:
                    # Create query URL
                    query_url = create_query_url(
                        website["url"], website["query"], product_code
                    )

                    # Get search page
                    active_browser.get(query_url)

                    # Get product URL from HTML
                    product_url = find_href_in_element(
                        active_browser,
                        website["product_element"],
                    )

                    # Add product URL to excel
                    sheet[
                        get_column_letter(website_index + WEBSITES_START_COL_NUMBER)
                        + str(product_index + PRODUCTS_START_ROW_NUMBER)
                    ] = product_url

                    # Save excel file
                    workbook.save(EXCEL_FILE_PATH)
                except Exception:
                    print(
                        f'Could not add "{product_code}" to excel file for "{website["url"]}" website.'
                    )
        except Exception:
            print(f'Unexpected error for "{website["url"]}".')

    # Close excel and browser
    workbook.close()
    active_browser.close()

    print("Completed.")


if __name__ == "__main__":
    main()

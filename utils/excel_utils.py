"Excel Utils"

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# pylint: disable=broad-except


def open_workbook(excel_file_path):
    "Opens the workbook and returns the object"

    return load_workbook(excel_file_path)


def close_workbook(workbook: Workbook):
    "Closes the workbook"

    workbook.close()


def save_workbook(workbook: Workbook, excel_file_path: str):
    "Saves the workbook"

    workbook.save(excel_file_path)


def get_column_data_from_excel_sheet(
    sheet: Worksheet,
    col_number: int,
    row_number_start: int,
    row_number_end: int,
):
    """
    Gets data from specified excel sheet,
    from specified column number, in specified row number range
    """

    return [
        cell.value
        for cell in sheet[get_column_letter(col_number)][
            row_number_start - 1 : row_number_end
        ]
    ]


def add_data_to_excel_sheet_column(
    sheet: Worksheet,
    col_number: int,
    row_number_start: int,
    data: [str],
):
    """
    Adds data to specified excel sheet,
    from specified column number, in specified row number range
    """

    for i, cell_data in enumerate(data):
        if cell_data:
            sheet[f"{get_column_letter(col_number)}{row_number_start + i}"] = cell_data
        else:
            sheet[f"{get_column_letter(col_number)}{row_number_start + i}"] = ""

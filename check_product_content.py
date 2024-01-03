"""
This is an content checker for bathroom products.

- It gets the product codes from excel file,
- finds the product in multiple platforms,
- and check the product content.
"""

# pylint: disable=broad-except
# pylint: disable=broad-exception-raised

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver import Chrome

from utils.hepsiburada_utils import (
    get_product_url_from_hepsiburada,
    login_to_hepsiburada,
)
from utils.trendyol_utils import (
    get_product_url_from_trendyol,
    login_to_trendyol,
)

# ----------------------------
# COFIGURATIONS
# ----------------------------
EXCEL_FILE_PATH = "/path/to/file.xlsx"
SHEET_NAME = "sheet_name"
PRODUCTS_START_ROW_NUMBER = 0
PRODUCT_CODES_COL_NUMBER = 0
PRODUCT_NAMES_COL_NUMBER = 0
WEBSITES = {
    "trendyol": {
        "username": "user",
        "password": "pswrd",
        "url_col_number": 0,
        "name_col_number": 0,
        "desc_col_number": 0,
        "image_col_number": 0,
    },
    "hepsiburada": {
        "username": "user",
        "password": "pswrd",
        "url_col_number": 0,
        "name_col_number": 0,
        "desc_col_number": 0,
        "image_col_number": 0,
    },
}


# ----------------------------
# MAIN
# ----------------------------
def main():
    "Main code"

    # Open excel file
    workbook = load_workbook(EXCEL_FILE_PATH)
    try:
        sheet = workbook[SHEET_NAME]
    except Exception:
        sheet = workbook.active

    # Get product codes
    product_codes = [
        cell.value
        for cell in sheet[get_column_letter(PRODUCT_CODES_COL_NUMBER)][
            PRODUCTS_START_ROW_NUMBER - 1 :
        ]
    ]

    # Create browser
    active_browser = Chrome()

    # Maximize the window
    active_browser.maximize_window()

    try:
        # Login to Trendyol
        login_to_trendyol(
            active_browser,
            WEBSITES["trendyol"]["username"],
            WEBSITES["trendyol"]["password"],
        )

        # Get URLs
        for product_index, product_code in enumerate(product_codes):
            try:
                # Get Trendyol product URL
                product_url = get_product_url_from_trendyol(
                    active_browser, product_code
                )

                # Raise exception if product URL was not found
                if not product_url:
                    raise Exception(
                        f"Could not found product on Trendyol: {product_code}"
                    )

                # Add product URL to excel
                sheet[
                    get_column_letter(
                        WEBSITES["trendyol"]["url_col_number"],
                    )
                    + str(product_index + PRODUCTS_START_ROW_NUMBER)
                ] = product_url
            except Exception as error:
                print(str(error))

        # Save excel file
        workbook.save(EXCEL_FILE_PATH)

        # Login to Hepsi Burada
        login_to_hepsiburada(
            active_browser,
            WEBSITES["hepsiburada"]["username"],
            WEBSITES["hepsiburada"]["password"],
        )

        # Get URLs
        for product_index, product_code in enumerate(product_codes):
            try:
                # Get Hepsi Burada product URL
                product_url = get_product_url_from_hepsiburada(
                    active_browser, product_code
                )

                # Raise exception if product URL was not found
                if not product_url:
                    raise Exception(
                        f"Could not found product on Hepsi Burada: {product_code}"
                    )

                # Add product URL to excel
                sheet[
                    get_column_letter(
                        WEBSITES["hepsiburada"]["url_col_number"],
                    )
                    + str(product_index + PRODUCTS_START_ROW_NUMBER)
                ] = product_url
            except Exception as error:
                print(str(error))

        # Save excel file
        workbook.save(EXCEL_FILE_PATH)

    except Exception:
        print("An unexpected error occured.")

    # Close excel and browser
    workbook.close()
    active_browser.close()

    print("Completed.")


if __name__ == "__main__":
    main()
